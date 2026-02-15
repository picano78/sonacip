"""
Data Export Utilities
Support for CSV, Excel, and PDF exports
"""
import csv
import io
from datetime import datetime, timezone
from flask import Response, make_response
import json


class DataExporter:
    """Handle data exports in various formats"""
    
    @staticmethod
    def to_csv(data, filename='export.csv', columns=None):
        """
        Export data to CSV
        
        Args:
            data: List of dictionaries or model objects
            filename: Output filename
            columns: List of column names (optional, auto-detected from first row)
            
        Returns:
            Flask Response with CSV file
        """
        # Convert model objects to dicts
        if data and hasattr(data[0], '__dict__'):
            data = [DataExporter._model_to_dict(obj) for obj in data]
        
        if not data:
            data = [{}]
        
        # Auto-detect columns
        if not columns:
            columns = list(data[0].keys())
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
        
        writer.writeheader()
        for row in data:
            writer.writerow(row)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
        return response
    
    @staticmethod
    def to_excel(data, filename='export.xlsx', sheet_name='Sheet1', columns=None):
        """
        Export data to Excel
        
        Args:
            data: List of dictionaries or model objects
            filename: Output filename
            sheet_name: Excel sheet name
            columns: List of column names
            
        Returns:
            Flask Response with Excel file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        # Convert model objects to dicts
        if data and hasattr(data[0], '__dict__'):
            data = [DataExporter._model_to_dict(obj) for obj in data]
        
        if not data:
            data = [{}]
        
        # Auto-detect columns
        if not columns:
            columns = list(data[0].keys())
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col_num, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, column in enumerate(columns, 1):
                value = row_data.get(column, '')
                # Handle datetime objects
                if isinstance(value, datetime):
                    value = value.isoformat()
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
        return response
    
    @staticmethod
    def to_json(data, filename='export.json', pretty=True):
        """
        Export data to JSON
        
        Args:
            data: List of dictionaries or model objects
            filename: Output filename
            pretty: Pretty print JSON
            
        Returns:
            Flask Response with JSON file
        """
        # Convert model objects to dicts
        if data and hasattr(data[0], '__dict__'):
            data = [DataExporter._model_to_dict(obj) for obj in data]
        
        # Serialize
        if pretty:
            json_str = json.dumps(data, indent=2, default=str)
        else:
            json_str = json.dumps(data, default=str)
        
        # Create response
        response = Response(
            json_str,
            mimetype='application/json; charset=utf-8',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
        return response
    
    @staticmethod
    def to_pdf(data, filename='export.pdf', title='Data Export', columns=None):
        """
        Export data to PDF
        
        Args:
            data: List of dictionaries or model objects
            filename: Output filename
            title: PDF title
            columns: List of column names
            
        Returns:
            Flask Response with PDF file
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        except ImportError:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        # Convert model objects to dicts
        if data and hasattr(data[0], '__dict__'):
            data = [DataExporter._model_to_dict(obj) for obj in data]
        
        if not data:
            data = [{}]
        
        # Auto-detect columns
        if not columns:
            columns = list(data[0].keys())
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Prepare table data
        table_data = [columns]
        for row in data:
            table_row = []
            for col in columns:
                value = row.get(col, '')
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M')
                table_row.append(str(value)[:50])  # Limit length
            table_data.append(table_row)
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        
        # Add footer
        elements.append(Spacer(1, 0.3 * inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        footer_text = f"Generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')} UTC"
        elements.append(Paragraph(footer_text, footer_style))
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        # Create response
        response = Response(
            output.getvalue(),
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename={filename}'
            }
        )
        
        return response
    
    @staticmethod
    def _model_to_dict(obj):
        """Convert SQLAlchemy model to dictionary"""
        result = {}
        for column in obj.__table__.columns:
            value = getattr(obj, column.name)
            # Handle datetime
            if isinstance(value, datetime):
                value = value.isoformat()
            result[column.name] = value
        return result
    
    @staticmethod
    def export_users(users, format='csv', filename=None):
        """Export users to file"""
        if not filename:
            filename = f'users_export_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.{format}'
        
        columns = ['id', 'username', 'email', 'first_name', 'last_name', 'role', 'is_active', 'created_at']
        
        data = []
        for user in users:
            data.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name or '',
                'last_name': user.last_name or '',
                'role': user.role.name if user.role else '',
                'is_active': user.is_active,
                'created_at': user.created_at.isoformat() if user.created_at else ''
            })
        
        if format == 'csv':
            return DataExporter.to_csv(data, filename, columns)
        elif format == 'excel':
            return DataExporter.to_excel(data, filename, 'Users', columns)
        elif format == 'json':
            return DataExporter.to_json(data, filename)
        elif format == 'pdf':
            return DataExporter.to_pdf(data, filename, 'Users Export', columns)
        else:
            raise ValueError(f"Unsupported format: {format}")
